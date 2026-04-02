import csv
import itertools
import json
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
INPUT_CSV = BASE_DIR / "slot_data_slorepo.csv"
OUTPUT_CSV = BASE_DIR / "slot_data_normalized.csv"
EXISTING_STORES = BASE_DIR / "existing_stores.txt"
STORE_MAPPING = BASE_DIR / "store_mapping.json"
EXISTING_MACHINES = BASE_DIR / "existing_machines.txt"
MACHINE_MAPPING = BASE_DIR / "machine_mapping.json"
UNMATCHED_STORES = BASE_DIR / "unmatched_stores.txt"
UNMATCHED_MACHINES = BASE_DIR / "unmatched_machines.txt"
MACHINE_MAPPING_XLSX = BASE_DIR / "機種名対応表.xlsx"

STORE_COLUMN = "店舗"
MACHINE_COLUMN = "機種名"
SIMILARITY_THRESHOLD = 0.70

STATUS_EXACT = "自動マッチ(完全一致)"
STATUS_SIMILAR = "自動マッチ(類似)"
STATUS_REVIEW = "要確認"
STATUS_NEW = "新機種"

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
STATUS_FILLS = {
    STATUS_EXACT: PatternFill(fill_type="solid", fgColor="FFFFFF"),
    STATUS_SIMILAR: PatternFill(fill_type="solid", fgColor="FFF2CC"),
    STATUS_REVIEW: PatternFill(fill_type="solid", fgColor="FCE5CD"),
    STATUS_NEW: PatternFill(fill_type="solid", fgColor="D9EAD3"),
}
DUPLICATE_HIGH_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
DUPLICATE_MEDIUM_FILL = PatternFill(fill_type="solid", fgColor="FCE5CD")
DUPLICATE_SIMILARITY_THRESHOLD = 0.80
DUPLICATE_HIGH_SIMILARITY_THRESHOLD = 0.95
MAX_DUPLICATE_CANDIDATES = 500


def normalize_name(name):
    return unicodedata.normalize("NFKC", name).replace(" ", "").replace("　", "")


def load_name_list(path):
    with path.open("r", encoding="utf-8") as file:
        return [line.strip() for line in file if line.strip()]


def load_mapping(path):
    if not path.exists():
        return {}

    with path.open("r", encoding="utf-8") as file:
        data = json.load(file)

    if not isinstance(data, dict):
        raise ValueError(f"{path.name} must contain a JSON object.")

    return data


def save_mapping(path, mapping):
    with path.open("w", encoding="utf-8") as file:
        json.dump(mapping, file, ensure_ascii=False, indent=2, sort_keys=True)
        file.write("\n")


def save_unmatched(path, names):
    with path.open("w", encoding="utf-8") as file:
        for name in sorted(names):
            file.write(f"{name}\n")


def build_normalized_lookup(existing_names):
    lookup = {}
    for existing_name in existing_names:
        lookup.setdefault(normalize_name(existing_name), existing_name)
    return lookup


def find_best_match(name, existing_names):
    best_name = None
    best_score = -1.0
    for existing_name in existing_names:
        score = SequenceMatcher(None, name, existing_name).ratio()
        if score > best_score:
            best_score = score
            best_name = existing_name
    return best_name, best_score


def similarity_score(left_name, right_name):
    if not left_name or not right_name:
        return None
    if normalize_name(left_name) == normalize_name(right_name):
        return 1.0
    return SequenceMatcher(None, left_name, right_name).ratio()


def build_machine_mapping_record(source_name, target_name, similarity, status):
    return {
        "slorepo機種名": source_name,
        "Ppro機種名(Supabase)": target_name,
        "類似度": similarity,
        "ステータス": status,
    }


def resolve_machine_status(source_name, target_name, similarity, exists_in_existing):
    if target_name and normalize_name(source_name) == normalize_name(target_name):
        return STATUS_EXACT
    if target_name and similarity >= SIMILARITY_THRESHOLD:
        return STATUS_SIMILAR
    if exists_in_existing:
        return STATUS_REVIEW
    return STATUS_NEW


def format_similarity(similarity):
    if similarity is None:
        return ""
    return round(similarity, 4)


def write_duplicate_candidate_sheet(workbook, records):
    worksheet = workbook.create_sheet("重複候補")
    headers = ["機種名A", "機種名B", "類似度", "NFKC正規化A", "NFKC正規化B"]
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for record in records:
        worksheet.append(
            [
                record["機種名A"],
                record["機種名B"],
                format_similarity(record["類似度"]),
                record["NFKC正規化A"],
                record["NFKC正規化B"],
            ]
        )

        row_fill = (
            DUPLICATE_HIGH_FILL
            if record["類似度"] >= DUPLICATE_HIGH_SIMILARITY_THRESHOLD
            else DUPLICATE_MEDIUM_FILL
        )
        for cell in worksheet[worksheet.max_row]:
            cell.fill = row_fill

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = max_length + 2


def write_machine_mapping_excel(path, records, duplicate_candidates):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "機種名対応表"

    headers = ["slorepo機種名", "Ppro機種名(Supabase)", "類似度", "ステータス"]
    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL

    for record in records:
        worksheet.append(
            [
                record["slorepo機種名"],
                record["Ppro機種名(Supabase)"],
                format_similarity(record["類似度"]),
                record["ステータス"],
            ]
        )
        status_cell = worksheet.cell(row=worksheet.max_row, column=4)
        status_cell.fill = STATUS_FILLS.get(record["ステータス"], STATUS_FILLS[STATUS_EXACT])

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            cell_value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(cell_value))
        worksheet.column_dimensions[column_letter].width = max_length + 2

    write_duplicate_candidate_sheet(workbook, duplicate_candidates)
    workbook.save(path)


def find_duplicate_machine_candidates(existing_machines, machine_records):
    new_machine_names = {
        record["slorepo機種名"]
        for record in machine_records.values()
        if not record["Ppro機種名(Supabase)"]
    }
    all_machine_names = list(dict.fromkeys([*existing_machines, *sorted(new_machine_names)]))

    candidates = []
    for left_name, right_name in itertools.combinations(all_machine_names, 2):
        if left_name == right_name:
            continue

        normalized_left = normalize_name(left_name)
        normalized_right = normalize_name(right_name)
        score = SequenceMatcher(None, normalized_left, normalized_right).ratio()
        if score < DUPLICATE_SIMILARITY_THRESHOLD:
            continue

        candidates.append(
            {
                "機種名A": left_name,
                "機種名B": right_name,
                "類似度": score,
                "NFKC正規化A": normalized_left,
                "NFKC正規化B": normalized_right,
            }
        )

    candidates.sort(key=lambda record: record["類似度"], reverse=True)
    return candidates[:MAX_DUPLICATE_CANDIDATES]


def normalize_value(name, existing_names, normalized_lookup, mapping, unmatched_names):
    if name in mapping:
        return mapping[name], "mapped"

    normalized = normalize_name(name)
    if normalized in normalized_lookup:
        canonical = normalized_lookup[normalized]
        if name != canonical:
            mapping[name] = canonical
            return canonical, "matched"
        return canonical, "unchanged"

    best_name, best_score = find_best_match(name, existing_names)
    if best_name is not None and best_score >= SIMILARITY_THRESHOLD:
        mapping[name] = best_name
        unmatched_names.add(name)
        return best_name, "matched"

    unmatched_names.add(name)
    return name, "unmatched"


def main():
    existing_stores = load_name_list(EXISTING_STORES)
    existing_machines = load_name_list(EXISTING_MACHINES)

    store_mapping = load_mapping(STORE_MAPPING)
    machine_mapping = load_mapping(MACHINE_MAPPING)

    store_lookup = build_normalized_lookup(existing_stores)
    machine_lookup = build_normalized_lookup(existing_machines)

    converted_store_count = 0
    converted_machine_count = 0
    unmatched_stores = set()
    unmatched_machines = set()
    machine_records = {}

    with INPUT_CSV.open("r", encoding="utf-8-sig", newline="") as infile:
        reader = csv.DictReader(infile)
        fieldnames = reader.fieldnames

        if not fieldnames:
            raise ValueError("slot_data_slorepo.csv is empty.")

        if STORE_COLUMN not in fieldnames or MACHINE_COLUMN not in fieldnames:
            raise ValueError(
                f"CSV must contain '{STORE_COLUMN}' and '{MACHINE_COLUMN}' columns."
            )

        rows = []
        for row in reader:
            original_store = row[STORE_COLUMN]
            original_machine = row[MACHINE_COLUMN]

            normalized_store, store_status = normalize_value(
                original_store,
                existing_stores,
                store_lookup,
                store_mapping,
                unmatched_stores,
            )
            normalized_machine, machine_status = normalize_value(
                original_machine,
                existing_machines,
                machine_lookup,
                machine_mapping,
                unmatched_machines,
            )

            if store_status in {"mapped", "matched"} and normalized_store != original_store:
                converted_store_count += 1
            if machine_status in {"mapped", "matched"} and normalized_machine != original_machine:
                converted_machine_count += 1

            best_machine_name, best_machine_score = find_best_match(
                original_machine,
                existing_machines,
            )
            machine_target = normalized_machine if normalized_machine in existing_machines else best_machine_name
            machine_similarity = similarity_score(original_machine, machine_target)
            machine_status_label = resolve_machine_status(
                original_machine,
                machine_target,
                machine_similarity,
                best_machine_name is not None,
            )
            if machine_status == "unmatched":
                machine_target = best_machine_name if best_machine_name is not None else ""
                machine_similarity = similarity_score(original_machine, machine_target)
                machine_status_label = resolve_machine_status(
                    original_machine,
                    machine_target,
                    machine_similarity,
                    best_machine_name is not None,
                )

            machine_records.setdefault(
                original_machine,
                build_machine_mapping_record(
                    original_machine,
                    machine_target,
                    machine_similarity,
                    machine_status_label,
                ),
            )

            row[STORE_COLUMN] = normalized_store
            row[MACHINE_COLUMN] = normalized_machine
            rows.append(row)

    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as outfile:
        writer = csv.DictWriter(outfile, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    save_mapping(STORE_MAPPING, store_mapping)
    save_mapping(MACHINE_MAPPING, machine_mapping)
    save_unmatched(UNMATCHED_STORES, unmatched_stores)
    save_unmatched(UNMATCHED_MACHINES, unmatched_machines)
    duplicate_candidates = find_duplicate_machine_candidates(existing_machines, machine_records)
    write_machine_mapping_excel(
        MACHINE_MAPPING_XLSX,
        sorted(machine_records.values(), key=lambda record: record["slorepo機種名"]),
        duplicate_candidates,
    )

    print(f"変換した店舗名件数: {converted_store_count}")
    print(f"変換した機種名件数: {converted_machine_count}")
    print(f"unmatched店舗名件数: {len(unmatched_stores)}")
    print(f"unmatched機種名件数: {len(unmatched_machines)}")
    if duplicate_candidates:
        print(
            f"⚠️ 機種名の重複候補が{len(duplicate_candidates)}件見つかりました。"
            '機種名対応表.xlsx の「重複候補」シートを確認してください。'
        )
    print("unmatched_machines.txt を確認して machine_mapping.json に追記してください")


if __name__ == "__main__":
    main()
